"""
Artefact 2 — Root Cause Analysis via DCDG Graph Propagation.

Loads deviation and residual data produced by Artefact 1, builds a Directed
Cause-Dependency Graph (DCDG) from the plant's AutomationML topology, and
propagates sensor alarm states through the graph to rank simulation parameters
by their estimated causal influence on the observed deviations.

For each test case the ranked parameter list is saved to a CSV file in
RCA_Results/. An optional interactive cutoff step lets the user filter
low-influence parameters before saving.

Usage:
    python Scripts/Artefact_2/root_cause_main.py

Configure AML file, test case paths, and output directories in Scripts/Config.py
before running.
"""

import sys
import os

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from Arroyo_DCDG_PLUT import DCDG_Class
from A2_Utilities import createNameList
import math
import time
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import Config


if __name__ == "__main__":
    t0 = time.time()
    aml_file = Config.AML_FILE

    dcdg = DCDG_Class()
    dcdg.generate_graphfromAML(aml_file)
    dcdg.generate_interactive_graph(output_file=Config.RCA_INITIAL_GRAPH_OUTPUT)

    num_vs = dcdg.g.vcount()
    num_es = dcdg.g.ecount()
    print(f"Number of vertices in the graph: {num_vs}")
    print(f"Number of edges in the graph: {num_es}")

    # Designations — order in which sensors are listed in the Excel sheet
    wb = load_workbook(filename=Config.RCA_DESIGNATIONS)
    ws = wb.active
    designation_dict = {}
    for row in list(ws.rows)[1:]:
        if str(row[1].value) != "None" and str(row[2].value) != "None":
            designation_dict[row[0].value] = [str(c.value) for c in row[1:]]

    # Columns 0, 1, 4 in the Excel sheet: sensor type, sensor tag, sensor suffix
    designation_list = createNameList(list(designation_dict.values()), [0, 1, 4])

    # Column indices for actuator values in the XMV workbook
    actuator_columns = [1, 2]

    datapath = Config.TESTCASE_DIRECTORY
    testcase_dir_list = Config.TESTCASE_DIR_LIST
    changed_parameters = Config.CHANGED_PARAMETERS
    number_of_actuators = Config.NUMBER_OF_ACTUATORS
    number_of_sensors = Config.NUMBER_OF_SENSORS

    for case in testcase_dir_list:
        t0 = time.time()
        print(f"Processing test case: {case}")

        wb_alarms    = load_workbook(filename=datapath + case + "/" + Config.DEVIATION_OUTPUT_FILE)
        wb_xmv       = load_workbook(filename=datapath + case + "/" + Config.ACTUATOR_OUTPUT_FILE)
        wb_residuals = load_workbook(filename=datapath + case + "/" + Config.RESIDUAL_OUTPUT_FILE)

        ws_alarms    = wb_alarms.active
        ws_xmv       = wb_xmv.active
        ws_residuals = wb_residuals.active

        case_number = int(case.split("_")[-1])
        xmv_dict       = {}
        alarm_dict     = {}
        deviation_dict = {}
        accumulated_combined_influence_matrix = None

        counter = 1
        for counter in range(1, ws_alarms.max_row + 1):
            print(f"Processing time step {counter}")

            for i in range(number_of_actuators):
                xmv_dict[designation_list[number_of_sensors + i].replace("_State", "")] = (
                    ws_xmv.cell(row=counter, column=actuator_columns[i]).value / 100
                )

            dcdg.update_valves(xmv_dict)

            for j in range(number_of_sensors):
                if "AIR" not in designation_list[j]:
                    alarm_dict[designation_list[j].replace("V", "YIC").replace("_", "_Measurement_")] = (
                        ws_alarms.cell(row=counter, column=j + 1).value
                    )

            dcdg.add_sensorlabels(alarm_dict)
            dcdg.set_alarmstates(alarm_dict)

            for k in range(number_of_sensors):
                if "AIR" not in designation_list[k]:
                    deviation_dict[designation_list[k].replace("V", "YIC").replace("_", "_Measurement_")] = (
                        ws_residuals.cell(row=counter, column=k + 1).value
                    )

            dcdg.set_deviations(deviation_dict)

            if not any(alarm_dict.values()):
                print(f"No alarms at time step {counter} — skipping")
                continue

            (
                combined_influence_matrix,
                parameter_influence_matrix,
                parameter_names_total,
                statevariable_names_total,
                parameter_names_with_influence,
                statevariable_names_with_influence,
            ) = dcdg.evaluate_distance_matrix(0.5, update_distance=True)

            # Exponential time-weighting: values early deviations higher
            weight_factor_for_time = math.exp(-counter * Config.RCA_DECAY_FACTOR)

            if accumulated_combined_influence_matrix is None:
                accumulated_combined_influence_matrix = combined_influence_matrix
            else:
                accumulated_combined_influence_matrix += combined_influence_matrix * weight_factor_for_time

            row_sums = [
                sum(accumulated_combined_influence_matrix[row, :])
                for row in range(accumulated_combined_influence_matrix.shape[0])
            ]
            sorted_indices = sorted(range(len(row_sums)), key=lambda k: abs(row_sums[k]), reverse=True)
            top_n_indices = sorted_indices[: Config.RCA_TOP_N_PARAMETERS]

            if counter >= Config.RCA_ABORT_TIMESTEP:
                break

        t1 = time.time()
        print(f"Test case {case} finished at time step {counter} in {t1 - t0:.2f} seconds")
        print("----------------------------------------------------------------------")
        print(f"Changed parameters: {changed_parameters.get(case, 'Unknown test case')}")
        print("----------------------------------------------------------------------")

        data = {
            "Parameter": [parameter_names_total[i] for i in top_n_indices],
            "Influence": [row_sums[i] for i in top_n_indices],
        }
        df = pd.DataFrame(data)
        df.to_csv(
            f"{Config.RCA_RESULTS_DIR}/parameter_influence_testcase_{case_number}.csv",
            index=False,
        )

        print("----------------------------------------------------------------------")
        for index in sorted_indices:
            print(f"Parameter: {parameter_names_total[index]}, Influence: {row_sums[index]}")
        print("----------------------------------------------------------------------")

        cutoff_y_n = input("Apply relative cutoff? (y/n): ")
        if cutoff_y_n.lower() == "y":
            print("Enter a relative cutoff value in percent (e.g. 30 for 30%).")
            print("All parameters whose |Influence| is at least this percentage of the maximum are kept.")
            cutoff_percent = float(input("Cutoff in percent: "))
            max_influence = df["Influence"].abs().max()
            cutoff_value = (cutoff_percent / 100.0) * max_influence
            df_filtered = df[df["Influence"].abs() >= cutoff_value]
            print("Filtered DataFrame:")
            print(df_filtered)
            output_path = f"{Config.RCA_RESULTS_DIR}/parameter_influence_testcase_{case_number}.csv"
            df_filtered.to_csv(output_path, index=False)
            print(f"Filtered DataFrame saved to {output_path}")

    t1 = time.time()
    print(f"Total execution time: {t1 - t0:.2f} seconds")
