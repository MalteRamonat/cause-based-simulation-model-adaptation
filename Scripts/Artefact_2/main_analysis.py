"""
Main entry point for the AVEDAS Root Cause Analysis system.

Usage:
    python main_analysis.py --aml path/to/file.aml [--parallel] [--export output_prefix]
"""

import argparse
import sys
import os
import json
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Optional

from components.analysis_engine import AnalysisEngine
from config.constants import ProcessingConstants, OutputConstants


class AvedasAnalysis:
    """Main class for the AVEDAS Root Cause Analysis system."""

    def __init__(self, max_processes: Optional[int] = None):
        """Initialize the analysis engine.

        Args:
            max_processes: Maximum number of worker processes for parallel computation.
        """
        self.engine = AnalysisEngine(max_processes)
        self.results = {}

    def run_analysis(self, aml_file: str,
                     valve_states: Optional[Dict[str, float]] = None,
                     alarm_states: Optional[Dict[str, int]] = None,
                     deviations: Optional[Dict[str, float]] = None,
                     use_parallel: bool = True) -> Dict[str, Any]:
        """Run a full Root Cause Analysis.

        Args:
            aml_file: Path to the AutomationML (.aml) file.
            valve_states: Valve positions (optional).
            alarm_states: Alarm states (optional).
            deviations: Sensor deviations (optional).
            use_parallel: Whether to use parallel matrix computation.

        Returns:
            Analysis results as a dictionary.
        """
        print("=" * 60)
        print("AVEDAS Root Cause Analysis - Neue modulare Version")
        print("=" * 60)
        
        try:
            # 1. Initialisierung von AML
            self.engine.initialize_from_aml(aml_file)
            
            # 2. System-State aktualisieren (falls Daten vorhanden)
            if valve_states or alarm_states or deviations:
                print("\nAktualisiere Systemzustand...")
                self.engine.update_system_state(
                    valve_states or {},
                    alarm_states or {},
                    deviations or {}
                )
            
            # 3. Matrizen berechnen
            print(f"\nBerechne Influence-Matrizen ({'parallel' if use_parallel else 'sequentiell'})...")
            matrices = self.engine.compute_influence_matrices(use_parallel)
            
            # 4. Ranking erstellen
            print("\nErstelle Parameter-Ranking...")
            ranking = self.engine.get_parameter_ranking()
            
            # 5. Zusammenfassung erstellen
            summary = self.engine.get_analysis_summary()
            
            # 6. Ergebnisse zusammenstellen
            self.results = {
                'summary': summary,
                'ranking': ranking,
                'matrices': {name: matrix.tolist() if hasattr(matrix, 'tolist') else matrix 
                           for name, matrix in matrices.items()},
                'aml_file': aml_file,
                'processing_mode': 'parallel' if use_parallel else 'sequential'
            }
            
            # 7. Ergebnisse ausgeben
            self._print_results()
            
            return self.results
            
        except Exception as e:
            print(f"Analysis error: {e}")
            sys.exit(1)
    
    def _print_results(self):
        """Print analysis results to the console."""
        summary = self.results['summary']
        ranking = self.results['ranking']

        print("\n" + "=" * 60)
        print("ANALYSIS RESULTS")
        print("=" * 60)

        print(f"\nGraph statistics:")
        print(f"  Vertices: {summary['graph_statistics']['vertices']}")
        print(f"  Edges:    {summary['graph_statistics']['edges']}")

        print(f"\nMatrix statistics:")
        print(f"  Active alarms:   {summary['matrix_statistics']['active_alarms']}")
        print(f"  Possible alarms: {summary['matrix_statistics']['possible_alarms']}")
        print(f"  Parameters:      {summary['matrix_statistics']['parameters']}")
        print(f"  Computation time: {summary['computation_time']:.2f} s")

        print(f"\nMatrix dimensions:")
        for name, shape in summary['matrix_shapes'].items():
            print(f"  {name}: {shape}")

        print(f"\nTop {min(10, len(ranking))} most influential parameters:")
        print("-" * 50)
        for i, (param_name, influence) in enumerate(ranking[:10], 1):
            print(f"  {i:2d}. {param_name:<30} {influence:>10.4f}")
        
    def export_results(self, output_prefix: str):
        """Export analysis results to CSV and Excel files.

        Args:
            output_prefix: Filename prefix for all output files.
        """
        print(f"\nExporting results with prefix '{output_prefix}'...")
        
        # 1. Matrizen als CSV exportieren
        self.engine.export_matrices(output_prefix)
        
        # 2. Ranking als CSV exportieren
        ranking_df = pd.DataFrame(
            self.results['ranking'], 
            columns=['Parameter', 'Influence']
        )
        ranking_file = f"{output_prefix}_parameter_ranking.csv"
        ranking_df.to_csv(ranking_file, index=False)
        print(f"Parameter-Ranking exportiert nach: {ranking_file}")
        
        # 3. Zusammenfassung als JSON exportieren
        summary_file = f"{output_prefix}_analysis_summary.json"
        with open(summary_file, 'w', encoding='utf-8') as f:
            # Erstelle eine JSON-serialisierbare Version der Zusammenfassung
            json_summary = {
                'summary': self.results['summary'],
                'top_ranking': self.results['ranking'][:20],
                'aml_file': self.results['aml_file'],
                'processing_mode': self.results['processing_mode']
            }
            json.dump(json_summary, f, indent=2, ensure_ascii=False)
        print(f"Analyse-Zusammenfassung exportiert nach: {summary_file}")
        
        # 4. Detaillierter Excel-Report (optional)
        try:
            self._export_excel_report(output_prefix)
        except ImportError:
            print("Excel export unavailable (openpyxl not installed)")
    
    def _export_excel_report(self, output_prefix: str):
        """Erstellt einen detaillierten Excel-Report."""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, Alignment
            
            # Excel-Datei erstellen
            wb = openpyxl.Workbook()
            
            # Zusammenfassung
            ws_summary = wb.active
            ws_summary.title = "Summary"

            summary_data = [
                ["AVEDAS Root Cause Analysis — Report", ""],
                ["", ""],
                ["AML file", self.results['aml_file']],
                ["Processing mode", self.results['processing_mode']],
                ["Computation time", f"{self.results['summary']['computation_time']:.2f} s"],
                ["", ""],
                ["Graph statistics", ""],
                ["Vertices", self.results['summary']['graph_statistics']['vertices']],
                ["Edges", self.results['summary']['graph_statistics']['edges']],
                ["", ""],
                ["Matrix statistics", ""],
                ["Active alarms", self.results['summary']['matrix_statistics']['active_alarms']],
                ["Possible alarms", self.results['summary']['matrix_statistics']['possible_alarms']],
                ["Parameters", self.results['summary']['matrix_statistics']['parameters']],
            ]
            
            for row_data in summary_data:
                ws_summary.append(row_data)
            
            # Titel formatieren
            ws_summary['A1'].font = Font(bold=True, size=14)
            ws_summary.merge_cells('A1:B1')
            
            # Parameter-Ranking
            ws_ranking = wb.create_sheet("Parameter-Ranking")
            ranking_df = pd.DataFrame(
                self.results['ranking'], 
                columns=['Parameter', 'Influence']
            )
            
            for r in dataframe_to_rows(ranking_df, index=False, header=True):
                ws_ranking.append(r)
            
            # Header formatieren
            for cell in ws_ranking[1]:
                cell.font = Font(bold=True)
            
            excel_file = f"{output_prefix}_detailed_report.xlsx"
            wb.save(excel_file)
            print(f"Excel report saved to: {excel_file}")

        except Exception as e:
            print(f"Excel export error: {e}")


def load_test_data() -> Dict[str, Any]:
    """Return example test data for valve states, alarm states, and deviations."""
    return {
        'valve_states': {
            'XV_101_State': 50.0,
            'XV_102_State': 75.0,
            'XV_103_State': 25.0
        },
        'alarm_states': {
            'YIC_101_Measurement': 1,
            'YIC_102_Measurement': 0,
            'YIC_103_Measurement': 2
        },
        'deviations': {
            'YIC_101_Measurement': -2.5,
            'YIC_102_Measurement': 0.8,
            'YIC_103_Measurement': -4.2
        }
    }


def main():
    """Entry point for command-line usage."""
    parser = argparse.ArgumentParser(
        description="AVEDAS Root Cause Analysis",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiele:
  python main_analysis.py --aml model.aml
  python main_analysis.py --aml model.aml --parallel --export results
  python main_analysis.py --aml model.aml --test-data --export analysis_results
        """
    )
    
    parser.add_argument(
        '--aml', 
        required=True,
        help='Pfad zur AML-Datei'
    )
    
    parser.add_argument(
        '--parallel', 
        action='store_true',
        help='Verwende parallele Matrixberechnung (Standard: True)'
    )
    
    parser.add_argument(
        '--sequential',
        action='store_true',
        help='Use sequential matrix computation instead of parallel'
    )

    parser.add_argument(
        '--export',
        type=str,
        help='Filename prefix for output files (exports results when provided)'
    )

    parser.add_argument(
        '--test-data',
        action='store_true',
        help='Load example test data for valves, alarms and deviations'
    )

    parser.add_argument(
        '--max-processes',
        type=int,
        help='Maximum number of parallel worker processes'
    )

    args = parser.parse_args()

    if not os.path.exists(args.aml):
        print(f"Error: AML file not found: {args.aml}")
        sys.exit(1)

    use_parallel = True
    if args.sequential:
        use_parallel = False
    elif args.parallel:
        use_parallel = True

    test_data = load_test_data() if args.test_data else {}

    analysis = AvedasAnalysis(max_processes=args.max_processes)

    results = analysis.run_analysis(
        aml_file=args.aml,
        valve_states=test_data.get('valve_states'),
        alarm_states=test_data.get('alarm_states'),
        deviations=test_data.get('deviations'),
        use_parallel=use_parallel
    )

    if args.export:
        analysis.export_results(args.export)

    print(f"\nAnalysis complete!")
    if args.export:
        print(f"Results exported with prefix '{args.export}'.")


if __name__ == "__main__":
    main()
